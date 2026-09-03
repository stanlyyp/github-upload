import csv
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
OUT = ROOT / "processed"
GROUPS = {
    "er": {"daily": ["er-data.xlsx", "er-data-2.xlsx"], "quality": ["embarrass.xlsx"], "stations": ["embarrass-station.xlsx", "er.txt"]},
    "gr": {"daily": ["gr-data.xlsx"], "quality": ["greenriver.xlsx"], "stations": ["greenriver-station.xlsx", "gr.txt"]},
    "sr": {"daily": ["sr-data.xlsx", "sr-data-2.xlsx"], "quality": ["south.xlsx"], "stations": ["south-station.xlsx", "sr.txt"]},
}

def text(value):
    return "" if value is None else str(value).strip()

def site_id(value):
    value = text(value).replace("USGS-", "")
    return "USGS-" + value if value else ""

def number(value):
    try:
        return float(str(value).replace(",", ""))
    except (TypeError, ValueError):
        return None

def average_rows(rows, keys):
    grouped = defaultdict(list)
    for row in rows:
        grouped[tuple(row[key] for key in keys)].append(row)
    result = []
    for key, values in sorted(grouped.items()):
        numeric = [row["value"] for row in values if row["value"] is not None]
        if not numeric:
            continue
        item = dict(zip(keys, key))
        item.update(value=sum(numeric) / len(numeric), record_count=len(numeric), unit=next((row["unit"] for row in values if row["unit"]), ""), source=";".join(sorted(set(row["source"] for row in values))))
        result.append(item)
    return result

def write_csv(path, rows, fields):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows({field: row.get(field, "") for field in fields} for row in rows)

def read_daily(path):
    rows = []
    workbook = load_workbook(path, read_only=True, data_only=True)
    for worksheet in workbook.worksheets:
        if worksheet.title == "Title" or "Explanation" in worksheet.title:
            continue
        worksheet.reset_dimensions()
        iterator = worksheet.iter_rows(values_only=True)
        header = next(iterator, ())
        columns = {index: re.search(r"-(\d{5})-", text(value)).group(1) for index, value in enumerate(header) if re.search(r"-(\d{5})-", text(value))}
        flow_column = next((index for index, code in columns.items() if code == "00060"), None)
        if flow_column is None:
            continue
        for row in iterator:
            date = text(row[0] if row else "")
            value = number(row[flow_column] if flow_column < len(row) else None)
            if re.match(r"^\d{4}-\d{2}-\d{2}$", date) and value is not None:
                rows.append({"site_id": site_id(worksheet.title), "date": date, "value": value * 0.028316846592, "unit": "m3/s", "source": path.name, "parameter_code": "00060", "parameter_name": "Discharge"})
    return rows

def quality_metric(characteristic):
    name = characteristic.lower()
    patterns = [("00665", "Total Phosphorus", ("phosphorus",)), ("62855", "Total Nitrogen", ("nitrogen, mixed", "total nitrogen")), ("80154", "Suspended Sediment Concentration", ("suspended sediment concentration",)), ("80155", "Suspended Sediment Discharge", ("suspended sediment discharge",)), ("70331", "Suspended Sediment Percent Composition", ("percent composition", "finer fraction")), ("00530", "Total Suspended Solids", ("total suspended solids",)), ("00535", "General Solids", ("general solids",))]
    for code, label, terms in patterns:
        if any(term in name for term in terms):
            return code, label
    return None

def read_quality(path):
    rows = []
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook["report"] if "report" in workbook.sheetnames else workbook.worksheets[0]
    worksheet.reset_dimensions()
    iterator = worksheet.iter_rows(values_only=True)
    header = [text(value) for value in next(iterator, ())]
    positions = {name: index for index, name in enumerate(header)}
    required = {"ActivityStartDate", "MonitoringLocationIdentifier", "CharacteristicName", "ResultMeasureValue"}
    if not required.issubset(positions):
        return rows
    for row in iterator:
        metric = quality_metric(text(row[positions["CharacteristicName"]]))
        date = text(row[positions["ActivityStartDate"]])
        value = number(row[positions["ResultMeasureValue"]])
        if not metric or value is None or not re.match(r"^\d{4}-\d{2}-\d{2}$", date):
            continue
        unit_index = positions.get("ResultMeasure/MeasureUnitCode")
        rows.append({"site_id": site_id(row[positions["MonitoringLocationIdentifier"]]), "date": date, "value": value, "unit": text(row[unit_index]) if unit_index is not None else "", "source": path.name, "parameter_code": metric[0], "parameter_name": metric[1]})
    return rows

def read_station_xlsx(path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook["report"] if "report" in workbook.sheetnames else workbook.worksheets[0]
    worksheet.reset_dimensions()
    iterator = worksheet.iter_rows(values_only=True)
    header = [text(value) for value in next(iterator, ())]
    positions = {name: index for index, name in enumerate(header)}
    rows = []
    for row in iterator:
        identifier_index = positions.get("MonitoringLocationIdentifier")
        if identifier_index is None or not text(row[identifier_index]):
            continue
        get = lambda name: text(row[positions[name]]) if name in positions else ""
        rows.append({"site_id": site_id(row[identifier_index]), "station_name": get("MonitoringLocationName"), "latitude": get("LatitudeMeasure"), "longitude": get("LongitudeMeasure"), "drainage_area": get("DrainageAreaMeasure/MeasureValue"), "state": "", "county": "", "data_source": path.name})
    return rows

def read_station_txt(path):
    lines = path.read_text(encoding="utf-8", errors="replace").splitlines()
    header_index = next((i for i, line in enumerate(lines) if line.startswith("agency_cd")), None)
    if header_index is None:
        return []
    headers = lines[header_index].split()
    rows = []
    for line in lines[header_index + 2:]:
        if not line.strip() or line.startswith("#"):
            continue
        fields = line.split(None, len(headers) - 1)
        if len(fields) >= 6:
            values = dict(zip(headers, fields))
            rows.append({"site_id": site_id(values.get("site_no")), "station_name": values.get("station_nm", ""), "latitude": values.get("dec_lat_va", ""), "longitude": values.get("dec_long_va", ""), "drainage_area": "", "state": "", "county": "", "data_source": path.name})
    return rows

def main():
    fields = ["site_id", "date", "value", "unit", "parameter_code", "parameter_name", "record_count", "source"]
    station_fields = ["site_id", "station_name", "latitude", "longitude", "drainage_area", "state", "county", "data_source"]
    OUT.mkdir(exist_ok=True)
    for group, files in GROUPS.items():
        flow = [row for name in files["daily"] if (ROOT / name).exists() for row in read_daily(ROOT / name)]
        quality = [row for name in files["quality"] if (ROOT / name).exists() for row in read_quality(ROOT / name)]
        flow = average_rows(flow, ["site_id", "date", "parameter_code", "parameter_name"])
        quality = average_rows(quality, ["site_id", "date", "parameter_code", "parameter_name"])
        write_csv(OUT / f"{group}_flow.csv", flow, fields)
        for code, name in (("00665", "tp"), ("62855", "tn")):
            write_csv(OUT / f"{group}_{name}.csv", [row for row in quality if row["parameter_code"] == code], fields)
        write_csv(OUT / f"{group}_tss.csv", [row for row in quality if row["parameter_code"] not in {"00665", "62855"}], fields)
        stations = []
        for name in files["stations"]:
            path = ROOT / name
            if path.exists():
                stations.extend(read_station_xlsx(path) if path.suffix == ".xlsx" else read_station_txt(path))
        station_by_id = {}
        for row in stations:
            station_by_id.setdefault(row["site_id"], row)
        write_csv(OUT / f"{group}_stations.csv", sorted(station_by_id.values(), key=lambda row: row["site_id"]), station_fields)
        print(group, "flow", len(flow), "tp", sum(row["parameter_code"] == "00665" for row in quality), "tn", sum(row["parameter_code"] == "62855" for row in quality), "tss", sum(row["parameter_code"] not in {"00665", "62855"} for row in quality), "stations", len(station_by_id))

if __name__ == "__main__":
    main()